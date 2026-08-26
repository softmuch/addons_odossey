import base64
import io

from odoo import models, fields, _
from odoo.exceptions import UserError


class PricelistBulkUpdateWizard(models.TransientModel):
    _name = 'pricelist.bulk.update.wizard'
    _description = 'Actualización Masiva de Precios de Lista'

    pricelist_id = fields.Many2one(
        'product.pricelist', string='Lista de precios', required=True)
    import_file = fields.Binary(string='Excel de Precios')
    import_filename = fields.Char(string='Nombre del Archivo')

    def _export_headers(self):
        return {
            'id': _('ID'),
            'code': _('Code'),
            'name': _('Name'),
            'price': _('Price'),
            'new_price': _('New Price'),
        }

    def _get_product_info(self, item):
        if item.product_id:
            return item.product_id.default_code, item.product_id.display_name
        if item.product_tmpl_id:
            return item.product_tmpl_id.default_code, item.product_tmpl_id.name
        return False, item.name

    def action_download_excel(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("La librería 'openpyxl' no está instalada en el servidor."))

        if not self.pricelist_id:
            raise UserError(_("Debes seleccionar una Lista de Precios."))

        items = self.env['product.pricelist.item'].search([('pricelist_id', '=', self.pricelist_id.id)])

        labels = self._export_headers()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Precios'
        ws.append([labels['id'], labels['code'], labels['name'], labels['price'], labels['new_price']])

        for item in items:
            default_code, name = self._get_product_info(item)
            ws.append([item.id, default_code or '', name or '', item.fixed_price, None])

        buffer = io.BytesIO()
        wb.save(buffer)

        attachment = self.env['ir.attachment'].create({
            'name': 'Precios_%s.xlsx' % self.pricelist_id.name,
            'type': 'binary',
            'datas': base64.b64encode(buffer.getvalue()),
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_import_excel(self):
        try:
            import openpyxl
        except ImportError:
            raise UserError(_("La librería 'openpyxl' no está instalada en el servidor."))

        if not self.pricelist_id:
            raise UserError(_("Debes seleccionar una Lista de Precios."))
        if not self.import_file:
            raise UserError(_("Debes adjuntar un archivo Excel para importar."))

        data = base64.b64decode(self.import_file)
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active

        labels = self._export_headers()

        headers = {}
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col).value
            if value is None:
                continue
            headers[str(value).strip()] = col

        missing = [labels[key] for key in ('id', 'new_price') if labels[key] not in headers]
        if missing:
            raise UserError(_("El archivo no contiene las columnas obligatorias: %s") % ', '.join(missing))

        id_col = headers[labels['id']]
        new_price_col = headers[labels['new_price']]

        errors = []
        skipped = 0
        rows_by_id = {}

        for row in range(2, ws.max_row + 1):
            item_id = ws.cell(row=row, column=id_col).value
            new_price = ws.cell(row=row, column=new_price_col).value

            if item_id is None:
                continue
            if new_price is None or (isinstance(new_price, str) and not new_price.strip()):
                skipped += 1
                continue
            if not isinstance(new_price, (int, float)):
                errors.append(
                    _("Fila %(row)s: '%(label)s' inválido (%(value)r), se omite.") % {
                        'row': row, 'label': labels['new_price'], 'value': new_price})
                continue
            try:
                item_id = int(item_id)
            except (TypeError, ValueError):
                errors.append(_("Fila %s: id inválido (%r), se omite.") % (row, item_id))
                continue

            rows_by_id[item_id] = new_price

        updated = 0
        if rows_by_id:
            items = self.env['product.pricelist.item'].browse(list(rows_by_id.keys())).exists()
            found_ids = set(items.ids)
            for item_id in rows_by_id:
                if item_id not in found_ids:
                    errors.append(_("Fila con id %s: no existe ningún Precio asociado, se omite.") % item_id)

            wrong_pricelist = items.filtered(lambda i: i.pricelist_id.id != self.pricelist_id.id)
            if wrong_pricelist:
                errors.append(
                    _("%s fila(s) pertenecen a otra Lista de Precios, se omiten.") % len(wrong_pricelist))
            items = items - wrong_pricelist

            for item in items:
                new_price = rows_by_id[item.id]
                if item.fixed_price != new_price:
                    item.fixed_price = new_price
                    updated += 1

        message = _("Precios actualizados: %(updated)s\nFilas sin '%(label)s': %(skipped)s") % {
            'updated': updated,
            'skipped': skipped,
            'label': labels['new_price'],
        }
        if errors:
            message += "\n\n" + "\n".join(errors[:50])

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Importación de Precios"),
                'message': message,
                'type': 'warning' if errors else 'success',
                'sticky': True,
            }
        }
